# -*- coding: utf-8 -*-
import os
import sys
import io
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from copy import copy
from datetime import datetime

from dotenv import load_dotenv

# Load biến môi trường từ .env
load_dotenv()

# ============================================================
# CẤU HÌNH HỆ THỐNG
# ============================================================
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

class QuotationConfig:
    TEMPLATE_PATH = os.getenv('TEMPLATE_PATH', 'Bao_Gia_Mau.xlsx')
    DINH_MUC_PATH = os.getenv('DINH_MUC_PATH', 'Dinh_Muc_Phan_Mem_Full.xlsx')
    LOGO_PATH = os.getenv('LOGO_PATH', os.path.join('images', 'logo.png'))
    OUTPUT_DIR = '.'
    
    MY_COMPANY = {
        "name": os.getenv('MY_COMPANY_NAME', "Công ty TNHH Anh Lập trình"),
        "address": os.getenv('MY_COMPANY_ADDRESS', ""),
        "phone": os.getenv('MY_COMPANY_PHONE', ""),
        "email": os.getenv('MY_COMPANY_EMAIL', ""),
        "bank_name": os.getenv('MY_COMPANY_BANK_NAME', ""),
        "bank_account": os.getenv('MY_COMPANY_BANK_ACCOUNT', ""),
        "bank_holder": os.getenv('MY_COMPANY_BANK_HOLDER', ""),
    }

# ============================================================
# THÔNG TIN ĐƠN HÀNG CỤ THỂ (ĐỂ TRỐNG THEO YÊU CẦU)
# ============================================================
PARTNER = {
    "name": "",
    "fullname": "",
    "address": "",
    "contact": "",
    "phone": "",
    "email": "",
}

REQUESTED_FEATURES = []

TERMS = {
    "payment_method": os.getenv('TERMS_PAYMENT', "Thanh toán chuyển khoản hoặc tiền mặt"),
    "delivery_time": os.getenv('TERMS_DELIVERY', "Giao hàng trong vòng 7–14 ngày sau khi đặt cọc"),
    "warranty": os.getenv('TERMS_WARRANTY', "Bảo hành 12 tháng"),
    "note": os.getenv('TERMS_NOTE', "Giá trên đã bao gồm thuế VAT 10%"),
    "vat_percent": int(os.getenv('TERMS_VAT_PERCENT', 10)),
}

# ============================================================
# LỚP XỬ LÝ CHÍNH
# ============================================================
class QuotationEngine:
    def __init__(self, config):
        self.config = config
        self.price_map = self._load_price_map()

    def _load_price_map(self):
        price_map = {}
        if not os.path.exists(self.config.DINH_MUC_PATH):
            return price_map
        wb = openpyxl.load_workbook(self.config.DINH_MUC_PATH, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and row[4]:
                name = str(row[2]).strip()
                price_map[name] = {"price": row[4], "difficulty": row[1], "hours": row[3]}
        wb.close()
        return price_map

    def _find_item_data(self, feature_name):
        if feature_name in self.price_map: return self.price_map[feature_name]
        for key, data in self.price_map.items():
            if feature_name.lower() in key.lower() or key.lower() in feature_name.lower():
                return data
        return None

    def copy_cell_style(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def _safe_insert_rows(self, ws, idx, amount):
        """Chèn dòng một cách an toàn, giữ nguyên các vùng gộp ô phía dưới."""
        if amount <= 0: return
        
        # 1. Lưu lại tất cả các vùng gộp ô
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 2. Xóa các vùng gộp ô hiện tại để tránh lỗi khi insert
        for m_range in merged_ranges:
            try: ws.unmerge_cells(str(m_range))
            except: pass
            
        # 3. Chèn dòng
        ws.insert_rows(idx, amount)
        
        # 4. Áp dụng lại các vùng gộp ô với tọa độ đã được dịch chuyển
        for m_range in merged_ranges:
            m_min_row, m_max_row = m_range.min_row, m_range.max_row
            m_min_col, m_max_col = m_range.min_col, m_range.max_col
            
            # Nếu vùng gộp nằm hoàn toàn dưới vị trí chèn
            if m_min_row >= idx:
                m_min_row += amount
                m_max_row += amount
            # Nếu vùng gộp bao trùm vị trí chèn (hiếm khi xảy ra ở Footer)
            elif m_max_row >= idx:
                m_max_row += amount
            
            # Gộp lại
            ws.merge_cells(start_row=m_min_row, end_row=m_max_row, 
                           start_column=m_min_col, end_column=m_max_col)

    def _detect_template_structure(self, ws):
        header_row = 11
        footer_row = 22
        for r in range(1, 60):
            for c in range(1, 10):
                val = str(ws.cell(row=r, column=c).value or "")
                if "STT" == val: header_row = r
                if "Tổng cộng" in val:
                    footer_row = r
                    break
        return header_row + 1, footer_row

    def generate(self, partner, features, terms):
        if not features: return
        print(f"--- Đang tạo báo giá cho: {partner.get('fullname', 'Khách hàng mới')} ---")
        
        items = []
        for feat in features:
            data = self._find_item_data(feat)
            if data: items.append({"name": feat, "price": data['price'], "unit": "Gói", "qty": 1})

        wb = openpyxl.load_workbook(self.config.TEMPLATE_PATH)
        ws = wb.active
        start_row, footer_orig = self._detect_template_structure(ws)
        
        default_rows = footer_orig - start_row
        extra_needed = max(0, len(items) - default_rows)

        # Sử dụng phương thức chèn dòng an toàn
        if extra_needed > 0:
            self._safe_insert_rows(ws, footer_orig, extra_needed)
            print(f"-> Đã chèn thêm {extra_needed} dòng và bảo toàn cấu trúc Merge.")

        # Copy style
        for r in range(start_row, start_row + max(len(items), default_rows)):
            for c in range(1, 11):
                source = ws.cell(row=start_row, column=c)
                target = ws.cell(row=r, column=c)
                if r > start_row: self.copy_cell_style(source, target)
                if r >= start_row + len(items): target.value = None

        # Điền dữ liệu
        for i, item in enumerate(items):
            r = start_row + i
            ws.cell(row=r, column=1).value = i + 1
            ws.cell(row=r, column=2).value = f"ALT-{i+1:03}"
            ws.cell(row=r, column=3).value = item['name']
            ws.cell(row=r, column=4).value = item['unit']
            ws.cell(row=r, column=5).value = item['qty']
            ws.cell(row=r, column=6).value = item['price']
            ws.cell(row=r, column=7).value = f"=E{r}*F{r}"

        # Cập nhật Footer
        f_start = footer_orig + extra_needed
        ws.cell(row=f_start, column=7).value = f"=SUM(G{start_row}:G{f_start-1})"
        ws.cell(row=f_start + 1, column=6).value = terms['vat_percent']
        
        # Điền các điều khoản (Vị trí tương đối từ f_start)
        # Lưu ý: Vì đã dùng _safe_insert_rows, các ô gộp trong Footer đã được bảo toàn
        ws.cell(row=f_start + 5, column=3).value = terms['payment_method']
        ws.cell(row=f_start + 6, column=3).value = terms['delivery_time']
        ws.cell(row=f_start + 7, column=3).value = terms['warranty']
        ws.cell(row=f_start + 8, column=3).value = terms['note']

        # Ngân hàng
        seller = self.config.MY_COMPANY
        ws.cell(row=f_start + 11, column=2).value = seller['bank_name']
        ws.cell(row=f_start + 12, column=2).value = seller['bank_account']
        ws.cell(row=f_start + 13, column=2).value = seller['bank_holder']

        # Lưu file
        p_name = partner.get('name', 'KhachHangMoi') or 'KhachHangMoi'
        filename = f"Bao_Gia_{p_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        print(f"THÀNH CÔNG: Đã lưu tại {filename}")

if __name__ == "__main__":
    engine = QuotationEngine(QuotationConfig)
    engine.generate(PARTNER, REQUESTED_FEATURES, TERMS)
